#!/mnt/sda1/opkg/usr/bin/python
#coding:utf-8
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from  openpyxl.reader.excel  import  load_workbook
from openpyxl.utils import get_column_letter
import typehelper
import time

class XlsHelper(object):
    def __init__(self,filename= r'../excel/test.xlsx'):
        self.filename = filename

    def GetSheetToDicts(self,sheetname):
        wb = load_workbook(self.filename)
        ws = wb.get_sheet_by_name(sheetname)
        header = map(lambda x:x.value,ws[1])
        records = self.GetRows(ws,2,ws.max_row,1,ws.max_column+1)
        dicts = typehelper.ListsToDicts(header,records)
        wb.close()
        print dicts
        return dicts

    def SaveDictsToSheet(self,sheetname,dicts):
        wb = load_workbook(self.filename)
        if sheetname  in wb.get_sheet_names():
            #wb.remove_sheet()
            sheetname+=time.strftime('%Y%m%d')
        wb.create_sheet(sheetname)
        ws = wb.get_sheet_by_name(sheetname)
        ks =sorted(dicts[0].keys())
        #header
        for i in range(1,len(ks)+1):
            ws.cell(row=1,column=i).value=ks[i-1]
        #body : Row or column values must be at least 1
        for j in range(2,len(dicts)+2):
            for i in range(1,len(ks)+1):
                ws.cell(row=j,column=i).value=dicts[j-2].get(ks[i-1])
        wb.save(self.filename)
        wb.close()

    def GetWorkSheet(self,sheetname):
        #取第一张表
        #sheetnames = self.wb.get_sheet_names()
        wb = load_workbook(self.filename)
        ws = wb.get_sheet_by_name(sheetname)
        wb.close()
        #显示表名，表行数，表列数
        print   "Work Sheet Titile:" ,ws.title
        print   "Work Sheet Rows:" ,ws.max_row
        print   "Work Sheet Cols:" ,ws.max_column
        return ws

    def GetRows(self,ws,rowstart,rowend,colstart,colend):
        records = []
        for rownum in range(rowstart,rowend+1):
            records.append(self.GetRow(ws,rownum,colstart,colend))
        return records

    def GetRow(self,ws,rownum,colstart,colend):
        record = []
        for colnum in range(colstart,colend):
            record.append(ws.cell(row=rownum,column=colnum).value)
        return record

    def WriteRows(self,ws,results,rowstart,rowend,colstart,colend):
        i=0
        for rownum in range(rowstart,rowend+1):
            self.WriteRow(ws,results[i],rownum,colstart,colend)
            i+=1
        ws.save(self.filename)

    def WriteRow(self,ws,result,rownum,colstart,colend):
        i=0
        print result
        for colnum in range(colstart,colend):
            ws.cell(row=rownum,column=colnum).value=result[i]
            i+=1

if __name__ == '__main__':
    xh=XlsHelper(filename= r'..\xlsx\sm.xlsx')
    ws = xh.GetWorkSheet(0)
    records= xh.GetRows(ws,2,16,1,10)
    print records[0]
